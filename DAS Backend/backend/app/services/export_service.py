"""
Export Service for Schedule Files
Supports Excel, PDF, and Image (PNG/JPG) export formats
"""

from typing import Dict, List, Any, Optional, BinaryIO
from datetime import datetime
from io import BytesIO
import os

# Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_TEXT
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# App Theme Colors (matching the frontend) - Modern & Clean
THEME_PRIMARY_BLUE = "3B82F6"      # Primary blue - hsl(211, 86%, 56%)
THEME_PRIMARY_DARK = "1E40AF"      # Deep blue for title
THEME_ACCENT_YELLOW = "FBBF24"     # Modern amber/gold - softer
THEME_SECONDARY_ORANGE = "F97316"  # Secondary orange - hsl(14, 90%, 62%)
THEME_LIGHT_BLUE = "F0F9FF"        # Very subtle blue tint
THEME_SOFT_BLUE = "DBEAFE"         # Soft blue for headers
THEME_LIGHT_GRAY = "F9FAFB"        # Almost white gray
THEME_BORDER_GRAY = "E5E7EB"       # Soft border color
THEME_TEXT_DARK = "1F2937"         # Dark text
THEME_TEXT_MEDIUM = "6B7280"       # Medium gray text
THEME_TEXT_LIGHT = "9CA3AF"        # Light gray text

# Modern clean font - Segoe UI is the modern Windows system font
# It's clean, contemporary, and renders Arabic beautifully
MODERN_FONT = "Segoe UI"

# PDF export
try:
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Image export
try:
    from PIL import Image, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

from sqlalchemy.orm import Session
from ..models.schedules import Schedule, ScheduleAssignment, TimeSlot
from ..models.academic import Subject, Class, AcademicYear
from ..models.teachers import Teacher


class ExportService:
    """Unified export service for all formats"""
    
    def __init__(self, db: Session):
        self.db = db
        self.day_names = ["", "الأحد", "الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت"]
    
    def export_schedule_excel(
        self,
        schedule_id: int,
        include_logo: bool = True,
        include_notes: bool = True
    ) -> BytesIO:
        """
        Export schedule to Excel format
        
        Args:
            schedule_id: Schedule ID
            include_logo: Whether to include school logo
            include_notes: Whether to include notes section
            
        Returns:
            BytesIO object with Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is not installed. Install it with: pip install openpyxl")
        
        # Get schedule data
        schedule_data = self._get_schedule_data(schedule_id)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "الجدول الدراسي"
        
        # Set document properties - Author as Educore
        wb.properties.creator = "Educore"
        wb.properties.lastModifiedBy = "Educore"
        wb.properties.title = "الجدول الدراسي"
        wb.properties.subject = "جدول الحصص الدراسية"
        wb.properties.company = "Educore"
        
        # Set RTL
        ws.sheet_view.rightToLeft = True
        
        # ============================================
        # MODERN STYLES - Clean & Contemporary
        # ============================================
        
        # Title style - Deep blue, clean modern look
        title_fill = PatternFill(start_color=THEME_PRIMARY_DARK, end_color=THEME_PRIMARY_DARK, fill_type="solid")
        title_font = Font(name=MODERN_FONT, bold=True, size=20, color="FFFFFF")
        
        # Header style - Primary Blue for day headers
        header_fill = PatternFill(start_color=THEME_PRIMARY_BLUE, end_color=THEME_PRIMARY_BLUE, fill_type="solid")
        header_font = Font(name=MODERN_FONT, bold=True, size=12, color="FFFFFF")
        
        # Period column style - Modern amber/gold
        period_fill = PatternFill(start_color=THEME_ACCENT_YELLOW, end_color=THEME_ACCENT_YELLOW, fill_type="solid")
        period_font = Font(name=MODERN_FONT, bold=True, size=11, color=THEME_TEXT_DARK)
        
        # Info section style - Soft blue background, clean text
        info_fill = PatternFill(start_color=THEME_SOFT_BLUE, end_color=THEME_SOFT_BLUE, fill_type="solid")
        info_font = Font(name=MODERN_FONT, bold=False, size=11, color=THEME_PRIMARY_DARK)
        
        # Cell styles for schedule content - clean and readable
        subject_font = Font(name=MODERN_FONT, bold=True, size=11, color=THEME_TEXT_DARK)
        teacher_font = Font(name=MODERN_FONT, size=10, color=THEME_TEXT_MEDIUM)
        cell_font = Font(name=MODERN_FONT, size=11, color=THEME_TEXT_DARK)
        empty_font = Font(name=MODERN_FONT, size=10, color=THEME_TEXT_LIGHT)
        
        # Alternating row colors - very subtle
        row_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        row_fill_alt = PatternFill(start_color=THEME_LIGHT_GRAY, end_color=THEME_LIGHT_GRAY, fill_type="solid")
        
        # Border styles - subtle and modern
        no_border = Border()  # No border for cleaner look
        subtle_border = Border(
            left=Side(style='thin', color=THEME_BORDER_GRAY),
            right=Side(style='thin', color=THEME_BORDER_GRAY),
            top=Side(style='thin', color=THEME_BORDER_GRAY),
            bottom=Side(style='thin', color=THEME_BORDER_GRAY)
        )
        bottom_border = Border(
            bottom=Side(style='thin', color=THEME_BORDER_GRAY)
        )
        
        # Alignments
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        # ============================================
        # TITLE SECTION - Modern clean header
        # ============================================
        row = 1
        ws.row_dimensions[row].height = 45
        ws.merge_cells(f'A{row}:F{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "الجدول الدراسي"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_alignment
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = title_fill
        row += 1
        
        # ============================================
        # INFO SECTION - Clean info bar
        # ============================================
        ws.row_dimensions[row].height = 32
        
        # Apply info styling to all cells first
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = info_fill
            cell.font = info_font
            cell.alignment = center_alignment
        
        # Class info (merged A:B)
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].value = f"الصف: {schedule_data['class_name']}"
        
        # Section info
        ws[f'C{row}'].value = f"الشعبة: {schedule_data['section'] or '1'}"
        
        # Session type
        ws[f'D{row}'].value = f"الفترة: {schedule_data['session_type']}"
        
        # Academic year (merged E:F)
        ws.merge_cells(f'E{row}:F{row}')
        ws[f'E{row}'].value = f"السنة الدراسية: {schedule_data['academic_year']}"
        
        # Small spacing row
        row += 1
        ws.row_dimensions[row].height = 6
        row += 1
        
        # ============================================
        # SCHEDULE TABLE - Modern clean design
        # ============================================
        periods = schedule_data['periods']
        days = [d for d in self.day_names if d]  # Exclude empty string
        
        # Header row - Clean blue headers
        ws.row_dimensions[row].height = 36
        
        # Period header cell
        period_header = ws[f'A{row}']
        period_header.value = "الحصة"
        period_header.fill = header_fill
        period_header.font = header_font
        period_header.alignment = center_alignment
        
        # Day header cells
        col_idx = 2
        for day in days[:5]:  # Sunday to Thursday
            cell = ws.cell(row=row, column=col_idx)
            cell.value = day
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            col_idx += 1
        
        # Data rows with subtle alternating colors
        row_index = 0
        for period_num in sorted(periods.keys()):
            row += 1
            row_index += 1
            ws.row_dimensions[row].height = 50
            
            # Period number with amber accent
            period_cell = ws[f'A{row}']
            period_cell.value = f"الحصة {period_num}"
            period_cell.fill = period_fill
            period_cell.font = period_font
            period_cell.alignment = center_alignment
            
            # Subtle alternating row color
            current_row_fill = row_fill_alt if row_index % 2 == 0 else row_fill_white
            
            # Subjects for each day
            col_idx = 2
            for day_idx in range(1, 6):  # Days 1-5 (Sunday-Thursday)
                cell = ws.cell(row=row, column=col_idx)
                
                # Find assignment for this day/period
                assignment = schedule_data['grid'].get((day_idx, period_num))
                if assignment:
                    # Format: Subject name on first line, teacher on second
                    cell.value = f"{assignment['subject']}\n{assignment['teacher']}"
                    cell.fill = current_row_fill
                    cell.font = cell_font
                else:
                    cell.value = "—"
                    cell.fill = current_row_fill
                    cell.font = empty_font
                
                cell.border = subtle_border
                cell.alignment = center_alignment
                col_idx += 1
        
        # ============================================
        # FOOTER - Clean timestamp
        # ============================================
        row += 1
        ws.row_dimensions[row].height = 25
        
        ws.merge_cells(f'A{row}:F{row}')
        timestamp_cell = ws[f'A{row}']
        timestamp_cell.value = f"Educore  |  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp_cell.alignment = Alignment(horizontal='center', vertical='center')
        timestamp_cell.font = Font(name=MODERN_FONT, size=9, color=THEME_TEXT_LIGHT)
        
        # ============================================
        # COLUMN WIDTHS - Clean proportions
        # ============================================
        ws.column_dimensions['A'].width = 14  # Period column
        for col_idx in range(2, 7):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        
        # Freeze header for scrolling
        ws.freeze_panes = 'A5'
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def export_schedule_pdf(
        self,
        schedule_id: int,
        orientation: str = "landscape",
        include_logo: bool = True
    ) -> BytesIO:
        """
        Export schedule to PDF format
        
        Args:
            schedule_id: Schedule ID
            orientation: portrait or landscape
            include_logo: Whether to include school logo
            
        Returns:
            BytesIO object with PDF file
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is not installed. Install it with: pip install reportlab")
        
        # Get schedule data
        schedule_data = self._get_schedule_data(schedule_id)
        
        # Create PDF
        output = BytesIO()
        page_size = landscape(A4) if orientation == "landscape" else A4
        doc = SimpleDocTemplate(output, pagesize=page_size, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        
        # Container for elements
        elements = []
        styles = getSampleStyleSheet()
        
        # Title style (RTL)
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        # Add title
        elements.append(Paragraph("الجدول الدراسي", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Add info
        info_style = ParagraphStyle('Info', parent=styles['Normal'], alignment=TA_RIGHT, fontSize=12)
        info_text = f"""
        <b>الصف:</b> {schedule_data['class_name']} &nbsp;&nbsp;&nbsp;
        <b>الشعبة:</b> {schedule_data['section'] or '1'} &nbsp;&nbsp;&nbsp;
        <b>الفترة:</b> {schedule_data['session_type']} &nbsp;&nbsp;&nbsp;
        <b>السنة:</b> {schedule_data['academic_year']}
        """
        elements.append(Paragraph(info_text, info_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Create table data
        periods = schedule_data['periods']
        days = [d for d in self.day_names[1:6]]  # Sunday to Thursday
        
        # Table header
        table_data = [['الحصة'] + days]
        
        # Table rows
        for period_num in sorted(periods.keys()):
            row = [f"الحصة {period_num}"]
            for day_idx in range(1, 6):
                assignment = schedule_data['grid'].get((day_idx, period_num))
                if assignment:
                    cell_text = f"{assignment['subject']}\n{assignment['teacher']}"
                else:
                    cell_text = "فارغ"
                row.append(cell_text)
            table_data.append(row)
        
        # Create table
        table = Table(table_data, repeatRows=1)
        
        # Table style
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data cells
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#D3D3D3')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (1, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
        ]))
        
        elements.append(table)
        
        # Add footer
        elements.append(Spacer(1, 0.3*inch))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], alignment=TA_CENTER, fontSize=9, textColor=colors.grey)
        elements.append(Paragraph(f"تم الإنشاء في: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        return output
    
    def export_schedule_image(
        self,
        schedule_id: int,
        format: str = "PNG",
        width: int = 1920,
        height: int = 1080
    ) -> BytesIO:
        """
        Export schedule to image format (PNG/JPG)
        
        Args:
            schedule_id: Schedule ID
            format: PNG or JPG
            width: Image width
            height: Image height
            
        Returns:
            BytesIO object with image file
        """
        if not PIL_AVAILABLE:
            raise ImportError("Pillow is not installed. Install it with: pip install Pillow")
        
        # Get schedule data
        schedule_data = self._get_schedule_data(schedule_id)
        
        # Create image
        img = Image.new('RGB', (width, height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Try to use a font (fallback to default if not available)
        try:
            title_font = ImageFont.truetype("arial.ttf", 40)
            header_font = ImageFont.truetype("arial.ttf", 28)
            cell_font = ImageFont.truetype("arial.ttf", 20)
            small_font = ImageFont.truetype("arial.ttf", 16)
        except:
            title_font = ImageFont.load_default()
            header_font = ImageFont.load_default()
            cell_font = ImageFont.load_default()
            small_font = ImageFont.load_default()
        
        # Colors
        header_bg = '#366092'
        header_text = 'white'
        cell_border = 'black'
        
        # Draw title
        title = "الجدول الدراسي"
        title_bbox = draw.textbbox((0, 0), title, font=title_font)
        title_width = title_bbox[2] - title_bbox[0]
        draw.text(((width - title_width) / 2, 40), title, fill='#366092', font=title_font)
        
        # Draw info
        info_y = 100
        info_text = f"الصف: {schedule_data['class_name']} | الشعبة: {schedule_data['section'] or '1'} | الفترة: {schedule_data['session_type']} | السنة: {schedule_data['academic_year']}"
        info_bbox = draw.textbbox((0, 0), info_text, font=small_font)
        info_width = info_bbox[2] - info_bbox[0]
        draw.text(((width - info_width) / 2, info_y), info_text, fill='black', font=small_font)
        
        # Calculate table dimensions
        periods = schedule_data['periods']
        num_periods = len(periods)
        num_days = 5
        
        table_start_y = 180
        table_width = width - 200
        table_height = height - table_start_y - 100
        
        cell_width = table_width // (num_days + 1)
        cell_height = table_height // (num_periods + 1)
        
        # Draw table header
        x = 100
        y = table_start_y
        
        # Header row
        draw.rectangle([x, y, x + cell_width, y + cell_height], fill=header_bg, outline=cell_border)
        header_text_bbox = draw.textbbox((0, 0), "الحصة", font=header_font)
        text_width = header_text_bbox[2] - header_text_bbox[0]
        text_height = header_text_bbox[3] - header_text_bbox[1]
        draw.text((x + (cell_width - text_width) / 2, y + (cell_height - text_height) / 2), 
                 "الحصة", fill=header_text, font=header_font)
        
        x += cell_width
        for day in self.day_names[1:6]:
            draw.rectangle([x, y, x + cell_width, y + cell_height], fill=header_bg, outline=cell_border)
            day_bbox = draw.textbbox((0, 0), day, font=header_font)
            text_width = day_bbox[2] - day_bbox[0]
            text_height = day_bbox[3] - day_bbox[1]
            draw.text((x + (cell_width - text_width) / 2, y + (cell_height - text_height) / 2), 
                     day, fill=header_text, font=header_font)
            x += cell_width
        
        # Data rows
        y += cell_height
        for period_num in sorted(periods.keys()):
            x = 100
            
            # Period number cell
            draw.rectangle([x, y, x + cell_width, y + cell_height], fill='#D3D3D3', outline=cell_border)
            period_text = f"الحصة {period_num}"
            period_bbox = draw.textbbox((0, 0), period_text, font=cell_font)
            text_width = period_bbox[2] - period_bbox[0]
            text_height = period_bbox[3] - period_bbox[1]
            draw.text((x + (cell_width - text_width) / 2, y + (cell_height - text_height) / 2), 
                     period_text, fill='black', font=cell_font)
            
            x += cell_width
            
            # Day cells
            for day_idx in range(1, 6):
                assignment = schedule_data['grid'].get((day_idx, period_num))
                draw.rectangle([x, y, x + cell_width, y + cell_height], fill='white', outline=cell_border)
                
                if assignment:
                    subject_text = assignment['subject']
                    teacher_text = assignment['teacher']
                    
                    # Draw subject
                    subject_bbox = draw.textbbox((0, 0), subject_text, font=cell_font)
                    text_width = subject_bbox[2] - subject_bbox[0]
                    draw.text((x + (cell_width - text_width) / 2, y + 10), 
                             subject_text, fill='black', font=cell_font)
                    
                    # Draw teacher
                    teacher_bbox = draw.textbbox((0, 0), teacher_text, font=small_font)
                    text_width = teacher_bbox[2] - teacher_bbox[0]
                    draw.text((x + (cell_width - text_width) / 2, y + cell_height - 30), 
                             teacher_text, fill='#666666', font=small_font)
                else:
                    empty_bbox = draw.textbbox((0, 0), "فارغ", font=cell_font)
                    text_width = empty_bbox[2] - empty_bbox[0]
                    text_height = empty_bbox[3] - empty_bbox[1]
                    draw.text((x + (cell_width - text_width) / 2, y + (cell_height - text_height) / 2), 
                             "فارغ", fill='#999999', font=cell_font)
                
                x += cell_width
            
            y += cell_height
        
        # Draw footer
        footer_text = f"تم الإنشاء في: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer_bbox = draw.textbbox((0, 0), footer_text, font=small_font)
        text_width = footer_bbox[2] - footer_bbox[0]
        draw.text(((width - text_width) / 2, height - 50), footer_text, fill='#999999', font=small_font)
        
        # Save to BytesIO
        output = BytesIO()
        img.save(output, format=format.upper(), quality=95 if format.upper() == 'JPEG' else None)
        output.seek(0)
        
        return output
    
    def _get_schedule_data(self, schedule_id: int) -> Dict[str, Any]:
        """Get schedule data for export"""
        schedule = self.db.query(Schedule).filter(Schedule.id == schedule_id).first()
        if not schedule:
            raise ValueError(f"Schedule with ID {schedule_id} not found")

        academic_year = self.db.query(AcademicYear).filter(
            AcademicYear.id == schedule.academic_year_id
        ).first()

        class_obj = self.db.query(Class).filter(Class.id == schedule.class_id).first()

        grade_level_ar = {
            "primary": "ابتدائي",
            "intermediate": "إعدادي",
            "secondary": "ثانوي"
        }

        class_name = (
            f"الصف {class_obj.grade_number} {grade_level_ar.get(class_obj.grade_level, '')}"
            if class_obj else "Unknown"
        )

        grid: Dict[tuple[int, int], Dict[str, str]] = {}
        periods: Dict[int, bool] = {}

        # In production data, each period may be stored directly as a Schedule row.
        related_schedules = self.db.query(Schedule).filter(
            Schedule.class_id == schedule.class_id,
            Schedule.section == schedule.section,
            Schedule.academic_year_id == schedule.academic_year_id,
            Schedule.session_type == schedule.session_type
        ).all()

        if related_schedules:
            subject_ids = {row.subject_id for row in related_schedules if row.subject_id}
            teacher_ids = {row.teacher_id for row in related_schedules if row.teacher_id}

            subjects_map = {
                subj.id: subj.subject_name
                for subj in self.db.query(Subject).filter(Subject.id.in_(subject_ids)).all()
            } if subject_ids else {}

            teachers_map = {
                teacher.id: teacher.full_name
                for teacher in self.db.query(Teacher).filter(Teacher.id.in_(teacher_ids)).all()
            } if teacher_ids else {}

            for row in related_schedules:
                if row.day_of_week is None or row.period_number is None:
                    continue

                grid[(row.day_of_week, row.period_number)] = {
                    'subject': subjects_map.get(row.subject_id, "Unknown"),
                    'teacher': teachers_map.get(row.teacher_id, "Unknown"),
                    'room': row.description or ""
                }

                periods[row.period_number] = True
        else:
            # Fallback to legacy ScheduleAssignment rows
            assignments = self.db.query(ScheduleAssignment).filter(
                ScheduleAssignment.schedule_id == schedule_id
            ).all()

            for assignment in assignments:
                time_slot = self.db.query(TimeSlot).filter(TimeSlot.id == assignment.time_slot_id).first()
                if not time_slot:
                    continue

                subject = self.db.query(Subject).filter(Subject.id == assignment.subject_id).first()
                teacher = self.db.query(Teacher).filter(Teacher.id == assignment.teacher_id).first()

                day = time_slot.day_of_week
                period = time_slot.period_number

                grid[(day, period)] = {
                    'subject': subject.subject_name if subject else "Unknown",
                    'teacher': teacher.full_name if teacher else "Unknown",
                    'room': assignment.room or ""
                }

                periods[period] = True

        return {
            'schedule_id': schedule_id,
            'class_name': class_name,
            'section': schedule.section,
            'session_type': "صباحي" if schedule.session_type == "morning" else "مسائي",
            'academic_year': academic_year.year_name if academic_year else "Unknown",
            'grid': grid,
            'periods': periods
        }
